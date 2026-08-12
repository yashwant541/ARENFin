"""Create Arabic sample spreadsheets that reflect real-world layouts.

Real Arabic financial workbooks store the label in **column A** and merely
*display* right-to-left (view mirrored). Those need NO column reversal — only
translation + switching the view to LTR. We also include one deliberately
mirrored file (label physically in the last column) to exercise the rare
reversal path.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import csv

OUT = Path(__file__).parent / "samples"
OUT.mkdir(exist_ok=True)
AR = Font(name="Arial", size=11)


def _style(ws, rtl_view=True):
    ws.sheet_view.rightToLeft = rtl_view
    for row in ws.iter_rows():
        for c in row:
            c.font = AR
            c.alignment = Alignment(horizontal="right")


# ---------------------------------------------------------------------------
# 1. Balance sheet — REAL Arabic convention.
#    Stored A..D = Label | Note | 2025 | 2024, sheet view = RTL (so column A
#    displays on the right). Correct output: keep column order, translate, LTR.
#    Expected: rtl_source=True, columns_reversed=False.
# ---------------------------------------------------------------------------
def balance_sheet_rtl():
    wb = Workbook(); ws = wb.active; ws.title = "الميزانية"
    rows = [
        ["قائمة المركز المالي الموحدة", "إيضاح", "٢٠٢٥", "٢٠٢٤"],
        ["الأصول", None, None, None],
        ["النقد والأرصدة لدى البنوك المركزية", "13", 77746, 63447],
        ["القروض والسلف للبنوك", "15", 43901, 43593],
        ["القروض والسلف للعملاء", "15", 286788, 281032],
        ["الاستثمارات في الأوراق المالية", None, 166956, 144556],
        ["الأدوات المالية المشتقة", "14", 65782, 81472],
        ["الشهرة والأصول غير الملموسة", "17", 6231, 5791],
        ["أصول أخرى", None, 67931, 43468],
        ["إجمالي الأصول", None, 919955, 849688],
        ["المطلوبات", None, None, None],
        ["ودائع البنوك", None, 30846, 25400],
        ["حسابات العملاء", None, 530161, 464489],
        ["سندات الدين المصدرة", "22", 72858, 64609],
        ["المطلوبات الثانوية", None, 8834, 10382],
        ["إجمالي المطلوبات", None, 865369, 798404],
        ["حقوق الملكية", None, None, None],
        ["رأس المال وعلاوة الإصدار", None, 6614, 6695],
        ["احتياطيات أخرى", None, 10406, 8724],
        ["الأرباح المبقاة", None, 29573, 28969],
        ["إجمالي حقوق الملكية", None, 54586, 51284],
        ["إجمالي حقوق الملكية والمطلوبات", None, 919955, 849688],
    ]
    for r in rows: ws.append(r)
    _style(ws, rtl_view=True)
    wb.save(OUT / "balance_sheet_ar.xlsx")


# ---------------------------------------------------------------------------
# 2. Income statement — REAL Arabic convention (label in column A, RTL view).
#    Expected: rtl_source=True, columns_reversed=False.
# ---------------------------------------------------------------------------
def income_statement_rtl():
    wb = Workbook(); ws = wb.active; ws.title = "قائمة الدخل"
    rows = [
        ["قائمة الدخل الموحدة", "إيضاح", "٢٠٢٥", "٢٠٢٤"],
        ["إيرادات الفوائد", None, 24547, 27862],
        ["مصروفات الفوائد", None, -18592, -21496],
        ["صافي إيرادات الفوائد", "3", 5955, 6366],
        ["صافي إيرادات الرسوم والعمولات", "4", 4249, 3734],
        ["صافي إيرادات المتاجرة", "5", 10294, 9615],
        ["إيرادات تشغيلية أخرى", "6", 444, -172],
        ["إجمالي الإيرادات التشغيلية", None, 20942, 19543],
        ["تكاليف الموظفين", None, -9109, -8510],
        ["المصروفات العمومية والإدارية", None, -2591, -2465],
        ["الاستهلاك والإطفاء", None, -1170, -1126],
        ["إجمالي المصروفات التشغيلية", "7", -13304, -12502],
        ["مخصص انخفاض قيمة الائتمان", "8", -672, -547],
        ["الربح قبل الضريبة", None, 6963, 6014],
        ["الضريبة", "10", -1866, -1972],
        ["ربح السنة", None, 5097, 4042],
        ["الحقوق غير المسيطرة", None, 12, -8],
        ["ربحية السهم الأساسية", "12", 195.4, 141.3],
    ]
    for r in rows: ws.append(r)
    _style(ws, rtl_view=True)
    wb.save(OUT / "income_statement_ar.xlsx")


# ---------------------------------------------------------------------------
# 3. Cash flow — "normal" Excel written in Arabic style (label in column A,
#    NO rtl view flag). Expected: rtl_source=True (Arabic present),
#    columns_reversed=False.
# ---------------------------------------------------------------------------
def cashflow_csv():
    rows = [
        ["قائمة التدفقات النقدية الموحدة", "٢٠٢٥", "٢٠٢٤"],
        ["الربح قبل الضريبه", 7638, 7041],                 # teh-marbuta variant
        ["الاستهلاك والاطفاء", 1170, 1126],
        ["صافي النقد من الانشطة التشغيليه", 9200, 8804],   # spelling drift -> fuzzy
        ["شراء الممتلكات والمعدات", -4100, -3200],
        ["النقد وما في حكمه في نهاية السنة", 77746, 63447],
        ["بند غير معروف تماماً", 620, 500],               # unknown -> untranslated
    ]
    with open(OUT / "cash_flow_ar.csv", "w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerows(rows)


# ---------------------------------------------------------------------------
# 4. Mirrored/visual storage — the RARE case that DOES need reversal.
#    Stored A..D = 2024 | 2025 | Note | Label (label physically in the LAST
#    column). Expected: rtl_source=True, columns_reversed=True.
# ---------------------------------------------------------------------------
def mirrored_visual():
    wb = Workbook(); ws = wb.active; ws.title = "ميزان مراجعة"
    rows = [
        ["٢٠٢٤", "٢٠٢٥", "إيضاح", "قائمة المركز المالي الموحدة"],
        [63447, 77746, "13", "النقد والأرصدة لدى البنوك المركزية"],
        [281032, 286788, "15", "القروض والسلف للعملاء"],
        [144556, 166956, None, "الاستثمارات في الأوراق المالية"],
        [849688, 919955, None, "إجمالي الأصول"],
        [464489, 530161, None, "حسابات العملاء"],
        [798404, 865369, None, "إجمالي المطلوبات"],
    ]
    for r in rows: ws.append(r)
    _style(ws, rtl_view=True)
    wb.save(OUT / "mirrored_visual_ar.xlsx")


if __name__ == "__main__":
    balance_sheet_rtl()
    income_statement_rtl()
    cashflow_csv()
    mirrored_visual()
    print("Wrote samples:", *[p.name for p in sorted(OUT.iterdir())])
